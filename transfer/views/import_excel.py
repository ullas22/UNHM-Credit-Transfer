from openpyxl import load_workbook
from django.shortcuts import render
from ..models.model_major import Major
from ..models.model_school import School
from ..models.model_course import Course
from ..models.model_major_requirement import Major_requirement
from ..models.model_transferevaluation import Transferevaluation
from ..models.model_approver import Approver

def import_file(request):
    if request.method == 'POST':
        file = request.FILES['document']
        import_data(file)
    return render(request, 'import.html')

def import_data(request):
    wb = load_workbook(filename=request, data_only=True)
    majors = import_major(wb)
    for major in majors:
        data = test_get_data_by_major(wb, major)
        schools = data[0]
        courses = data[1]
        approvers = data[2]
        reqs = data[3]
        evals = data[4]
        import_school(schools)
        import_course(courses)
        import_approvers(approvers)
        import_requirement(reqs)
        import_evaluations(evals)

def import_major(wb_object):
    '''
    Goes through the majors and adds them into the database
    '''
    major_names = wb_object.sheetnames
    count = 1
    for major in enumerate(major_names):
        major_data = Major(count, major[1])
        major_data.save()
        count = count + 1
    return major_names

def import_school(schools):
    '''
    Goes through the Schools and adds them into the database
    '''
    count = 1
    for school in schools:
        school_data = School(count, school, "N/A")
        school_data.save()
        count = count + 1
    return schools

def import_course(courses):
    '''
    Goes through the courses and adds them into the database
    '''
    count = 1
    for course in courses:
        course_data = Course(count, course[0], course[1], course[2])
        course_data.save()
        count = count + 1


def import_approvers(approvers):
    '''
    Goes through the approvers and adds them into the database
    '''
    count = 1
    for approver in approvers:
        approver_data = Approver(count, approver)
        approver_data.save()
        count = count + 1


def import_requirement(reqs):
    '''
    Goes through the requirements and adds them into the database
    '''
    count = 1
    for req in reqs:
        req_data = Major_requirement(count, req[1], req[0])
        req_data.save()
        count = count + 1


def import_evaluations(evals):
    '''
    Goes through the evaluations and adds them into the database
    '''
    count = 1
    for evaluation in evals:
        eval_data = Transferevaluation(count, evaluation[0], evaluation[1],
                                       evaluation[2], "2020-03-03",
                                       evaluation[3], evaluation[5],
                                       evaluation[4])
        eval_data.save()
        count = count + 1

def test_get_data_by_major(transfer_wb, major_name):
    """
    Tests get_data_by_major()
    """
    schools = []
    courses = []
    approvers = []
    major_reqs = []
    evals = []

    major_ws = transfer_wb[major_name]
    major_id = transfer_wb.sheetnames.index(major_name) + 1
    get_data_by_major(
        major_ws, major_id, schools, courses, approvers, major_reqs, evals
    )
    return [schools, courses, approvers, major_reqs, evals]



def get_unique_vals_from_col(major_ws, col_idx):
    """
    Computes and returns a list of unique values from the cell values of the
    column at col_idx in the worksheet major_ws.
    Used to get School and Approver name attribute values and
    MajorRequirement description attribute values.
    major_ws: worksheet
    col_idx: integer, column index
    Returns: tuple of unique values
    """
    # Accumulator is a set because it let us add ONLY unique values
    value_set = set()
    # Iterate over all the rows in the selected column, col_idx
    # Loop variable is a tuple with one element, call value
    for row_tuple in major_ws.iter_rows(
            min_row=2, max_row=major_ws.max_row,
            min_col=col_idx, max_col=col_idx,
            values_only=True):
        # Extract element value from the tuple and add it to value_set
        # If element value already exists in value_set, nothing happens
        value_set.add(row_tuple[0])
    return list(value_set)


def eval_with_fk(major_id, eval_row, schools, courses, approvers, major_reqs):
    """
    Uses data from eval_row (12-element list) and IDs from courses, approvers,
    and major_reqs to create and return a list that has foreign key values for
    approver ID, course ID, and major requirement ID.
    major_id: index of worksheet in the workbook incremented by 1
    eval_row: list of values from columns 1 to 12 in a row in a worksheet
    schools, courses, approvers, major_reqs: lists of data
    Returns: list of values representing an eval netry in the
    TransferEvaluation entity with forign keys
        course ID, majore requirement ID, and approver ID
    """
    eval_data = []

    # get school name from column 1
    # find school ID in schools list
    school_name = eval_row[0]
    school_id = schools.index(school_name) + 1

    # get subject number and course title from columns 2 and 3
    # find course ID in courses list
    course_data = []
    course_data.append(school_id)
    course_data.extend(eval_row[1:3])
    course_id = courses.index(course_data) + 1
    eval_data.append(course_id)

    # get major requirement description from column 6
    # find major requirement ID in major_reqs list
    major_req_desc = eval_row[5]
    major_req_id = major_reqs.index([major_id, major_req_desc]) + 1
    eval_data.append(major_req_id)

    # Get sem & year taken and approved status from columns 4 and 5
    eval_data.extend(eval_row[3:5])

    # get approver name from column 8 and find approver ID in approvers list
    approver_name = eval_row[7]
    approver_id = approvers.index(approver_name) + 1
    eval_data.append(approver_id)

    # get expiration date from column 9
    eval_data.append(eval_row[8])

    # get comment from column 12
    eval_data.append(eval_row[11])

    return eval_data


def course_with_fk(course_row, schools):
    """
    Uses data from course_row and the ID of of the school name  to create and
    return a list that has the foreign key value for school_id
    course_row: list of values from columns 1 to 3 in a row in a worksheet
    schools: list of strings, representing school names
    Returns:
        list of values representing a course entry in the Course entity with
        foreign key school ID
    """
    course_data = []
    school_name = course_row[0]
    school_id = schools.index(school_name) + 1
    course_data.append(school_id)
    course_data.extend(course_row[1:3])
    return course_data


def get_course_by_major(major_ws, start, end, schools):
    """
    Computes and returns a list of 3-element sublists with course data from
    major_ws worksheet. The 3-element sublist has data from columns
        1 (school name), 2 (subject num), and 3 (course title) in a worksheet.
    Replaces school name with school ID in the schools list.
    major_ws: worksheet
    start, end: column indices
    schools: lists of unique names
    Returns: list of 3-elemnt sublists
    """
    courses = []
    for row in major_ws.iter_rows(
            min_row=2, max_row=major_ws.max_row, min_col=start, max_col=end,
            values_only=True):
        course_data = course_with_fk(row, schools)
        courses.append(course_data)
    return courses


def get_eval_by_major(
        major_ws, major_id, start, end, schools, courses, approvers,
        major_reqs):
    """
    Computes and returns a list of 7-elemnent sublists with transfer evaluation
    data from major_ws worksheet. The 7-element sublist has data from
        4 columns:
            4 (sem & yea taken), 5 (approved status), 8 (approver name),
            9 (expiration date), 12 (comment)
        column 8 (approver name) is replaced with approver ID in approvers
        course ID from courses and major requirement ID from major_reqs
    majore_ws: worksheet
    major_id: index of worksheet in the workbook incremented by 1
    """
    evals = []
    for row in major_ws.iter_rows(
            min_row=2, max_row=major_ws.max_row, min_col=start, max_col=end,
            values_only=True):
        eval_data = eval_with_fk(
            major_id, row, schools, courses, approvers, major_reqs)
        evals.append(eval_data)
    return evals


def get_data_by_major(
        major_ws, major_id, schools, courses, approvers, major_reqs, evals):
    """
    Computes schools list with unique school names from major_ws
    Computes courses list with unique data composed of
        3-element sublists that have:
            school ID, subject number, and course title
    Computes approvers list with unique approver names from major_ws
    Computes major-reqs list with unique data composed of
        2-element sublists that have major ID and description
    Computes transfer evalaution list with unique data composed of
        2-element sublists that have:
            course ID and major requirement ID
    major_ws: worksheet
    major_id: index of worksheet in the workbook incremented by 1
    schools,  courses, approvers, major_reqs, evals:
        Empty lists are passed into the call and extended with data from
        the worksheet.
    Returns:
        nothing. We use pass by reference mutable objects to update the
        parameters
    """
    school_col_idx = 1  # columh 'A'
    school_lst = get_unique_vals_from_col(major_ws, school_col_idx)
    schools.extend(school_lst)

    start_col = 1
    end_col = 3
    course_lst = get_course_by_major(major_ws, start_col, end_col, schools)
    courses.extend(course_lst)

    approver_col_idx = 8  # column 'H'
    approver_lst = get_unique_vals_from_col(major_ws, approver_col_idx)
    approvers.extend(approver_lst)

    major_req_col_idx = 6  # column 'F'
    major_descriptions = get_unique_vals_from_col(major_ws, major_req_col_idx)
    for description in major_descriptions:
        req_data = [major_id, description]
        major_reqs.append(req_data)

    start_col = 1
    end_col = 12
    transfer_eval_lst = get_eval_by_major(
        major_ws, major_id, start_col, end_col, schools, courses, approvers,
        major_reqs)
    evals.extend(transfer_eval_lst)

def update(all_lst, one_lst):
    """
    Adds elements in one_lst to all_lst only if they are unique
    """
    for val in one_lst:
        if val not in all_lst:
            all_lst.append(val)
    return all_lst


def get_all_data(
        transfer_wb, all_majors, all_schools, all_courses, all_approvers,
        all_major_reqs, all_evals):
    """
    Builds all_majors list of major names
    Builds all_scholls list of school names
    Builds all_approvers list of names
    Builds all_major_reqs list of 2-element sublists composed of:
        major ID and description
    Builds all_courses list of 3-element sublists composed of:
        school ID, subject_num, and title
    Builds all_evals list of 7-element sublists composed of:
        course Id, majore requirement ID, semester/year taken, approved status,
        approver ID, expiration date, and comment
    transfer_wb: workbook
    all_majors, all_schools, all_approvers, all_major_reqs, all_courses,
    all_evals: lists, passed in as empty and built to have all the data for the
    database
    """
    # Iterate through all_major worksheets to get access to data on each
    # worksheet
    for idx, major_name in enumerate(all_majors):
        major_ws = transfer_wb[major_name]
        major_id = idx + 1
        schools = []
        approvers = []
        major_reqs = []
        courses = []
        evals = []
        # Passing empty lists and getting back lists with data for each entity
        get_data_by_major(
            major_ws, major_id, schools, courses, approvers, major_reqs, evals
        )
        # Use the worksheet lists to update build workbook lists
        update(all_schools, schools)
        update(all_approvers, approvers)
        update(all_major_reqs, major_reqs)
        update(all_courses, courses)
        update(all_evals, evals)
